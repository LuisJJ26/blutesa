import datetime
import os
import re

from django.core.management.base import BaseCommand, CommandError

from clientes.models import ClienteFibra

RUTA_POR_DEFECTO = os.path.expanduser('~/Downloads/CLIENTES INTERNET.xlsx')

# (nombre de hoja en el Excel, tipo en el modelo)
HOJAS = [
    ('INTERNET Y CABLE', ClienteFibra.Tipo.CABLE),
    ('INTERNET 5 G', ClienteFibra.Tipo.INALAMBRICO_5G),
]

PATRON_COSTO = re.compile(r'[\d.,]+')


def limpiar_costo(valor):
    """Devuelve (costo_decimal_o_None, nota). 'SOCIO' y similares van a la nota,
    no al monto — no son un precio en córdobas."""
    if valor is None:
        return None, ''
    if isinstance(valor, (int, float)):
        return round(float(valor), 2), ''

    texto = str(valor).strip()
    coincidencia = PATRON_COSTO.search(texto)
    if not coincidencia or not any(ch.isdigit() for ch in coincidencia.group()):
        return None, texto[:50]

    numero = coincidencia.group().replace(',', '')
    try:
        return round(float(numero), 2), ''
    except ValueError:
        return None, texto[:50]


def limpiar_fecha(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor

    solo_digitos = re.sub(r'\D', '', str(valor))
    if len(solo_digitos) == 8:
        dia, mes, anio = solo_digitos[:2], solo_digitos[2:4], solo_digitos[4:]
        try:
            return datetime.date(int(anio), int(mes), int(dia))
        except ValueError:
            return None
    return None


def limpiar_entero(valor):
    if isinstance(valor, (int, float)):
        return int(valor)
    if isinstance(valor, str) and valor.strip().isdigit():
        return int(valor.strip())
    return None


class Command(BaseCommand):
    help = (
        'Importa (o actualiza) el maestro de clientes de fibra/internet desde '
        'CLIENTES INTERNET.xlsx (hojas "INTERNET Y CABLE" e "INTERNET 5 G"). No incluye el '
        'historial de pagos mensuales del Excel, solo datos del cliente y del equipo instalado. '
        'Uso: importar_clientes_fibra_excel [--archivo RUTA]'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo', default=RUTA_POR_DEFECTO,
            help=f'Ruta al .xlsx (por defecto: {RUTA_POR_DEFECTO})',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Falta la librería openpyxl. Instálala con: pip install openpyxl')

        ruta = options['archivo']
        if not os.path.isfile(ruta):
            raise CommandError(f'No se encontró el archivo: {ruta}')

        libro = openpyxl.load_workbook(ruta, data_only=True, read_only=True)

        creados = 0
        actualizados = 0
        ips_existentes = set(ClienteFibra.objects.values_list('tipo', 'ip'))
        lote = []

        for nombre_hoja, tipo in HOJAS:
            if nombre_hoja not in libro.sheetnames:
                self.stdout.write(self.style.WARNING(f'Hoja "{nombre_hoja}" no encontrada, se omite.'))
                continue
            hoja = libro[nombre_hoja]

            encabezados = None
            for fila in hoja.iter_rows(values_only=True):
                if fila and any(str(c).strip() == 'Cliente' for c in fila if c is not None):
                    encabezados = fila
                    break
            if encabezados is None:
                self.stdout.write(self.style.WARNING(f'No se encontró el encabezado en "{nombre_hoja}", se omite.'))
                continue

            indice_encabezado = list(hoja.iter_rows(values_only=True)).index(encabezados) + 1

            for fila in hoja.iter_rows(min_row=indice_encabezado + 1, values_only=True):
                datos = dict(zip(encabezados, fila))
                nombre = (datos.get('Cliente') or '').strip() if datos.get('Cliente') else ''
                if not nombre or 'MIKROTIK' in nombre.upper():
                    continue

                ip = (datos.get('Total de BW') or '').strip() if datos.get('Total de BW') else ''
                costo, costo_nota = limpiar_costo(datos.get('Costo'))

                lote.append(ClienteFibra(
                    tipo=tipo,
                    ip=ip or None,
                    nombre=nombre,
                    cedula=(datos.get('Cedula') or '').strip() if datos.get('Cedula') else '',
                    barrio=(datos.get('BARRIO') or '').strip() if datos.get('BARRIO') else '',
                    telefono=str(datos.get('TELEFONO') or '').strip(),
                    pon=limpiar_entero(datos.get('PON')),
                    onu_id=limpiar_entero(datos.get('ID')),
                    serie=(datos.get('SN') or '').strip() if datos.get('SN') else '',
                    modelo=(datos.get('MODELO') or '').strip() if datos.get('MODELO') else '',
                    bw_mbps=limpiar_entero(datos.get('BW')),
                    costo=costo,
                    costo_nota=costo_nota,
                    fecha_activacion=limpiar_fecha(datos.get('Fecha de Activacion')),
                ))

                if (tipo, ip or None) in ips_existentes:
                    actualizados += 1
                else:
                    creados += 1

        if lote:
            ClienteFibra.objects.bulk_create(
                lote,
                update_conflicts=True,
                unique_fields=['tipo', 'ip'],
                update_fields=[
                    'nombre', 'cedula', 'barrio', 'telefono', 'pon', 'onu_id', 'serie',
                    'modelo', 'bw_mbps', 'costo', 'costo_nota', 'fecha_activacion',
                ],
            )

        self.stdout.write(self.style.SUCCESS(
            f'Importación completa: {creados} clientes nuevos, {actualizados} actualizados.'
        ))
